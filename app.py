"""
app.py — Flask backend Ágiles Azteca (Google Sheets).

Uso local:
    cp .env.example .env  →  completa SHEET_ID y GOOGLE_CREDENTIALS_JSON
    pip install -r requirements.txt
    python app.py           →  http://127.0.0.1:5050

Render.com:
    Start command: gunicorn app:app --workers 2 --timeout 120
    Configura las env vars en el panel de Render.
"""
import os
import json
import csv
import io
import re
import time
import threading
import unicodedata
from datetime import datetime, date, timedelta, timezone

import gspread
from google.oauth2.service_account import Credentials
from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ── Config ────────────────────────────────────────────────────────────
def _static_dir():
    """Sirve desde static/ si existe (local), o desde la raíz (Render sin carpeta static/)."""
    d = os.path.join(os.path.dirname(__file__), 'static')
    return d if os.path.isdir(d) else os.path.dirname(__file__)

app = Flask(__name__, static_folder=_static_dir())
CORS(app)

@app.errorhandler(gspread.exceptions.APIError)
def handle_sheets_api_error(e):
    """Maneja errores de Google Sheets API (especialmente 429 rate-limit) devolviendo caché o datos vacíos."""
    status = getattr(getattr(e, 'response', None), 'status_code', 0)
    carril_key = request.args.get('carril', 'mff').strip().lower() if request else 'mff'
    if status == 429:
        cached = _PORTAFOLIO_CACHE.get(carril_key)
        if cached:
            app.logger.warning(f'[portafolio] Rate-limit 429 — sirviendo caché para {carril_key}')
            return jsonify(cached['data'])
        app.logger.warning(f'[portafolio] Rate-limit 429 sin caché para {carril_key}')
    else:
        app.logger.error(f'[portafolio] Sheets APIError {status} para {carril_key}: {e}')
    return jsonify({
        'ultima_actualizacion': datetime.now(timezone.utc).isoformat(),
        'carriles': {}, 'proyectos': {}, 'portafolios': {},
        'liberaciones': [], 'hc_roster': [],
    })

SHEET_ID                = os.environ.get('SHEET_ID', '')
MASTER_SPREADSHEET_ID   = os.environ.get('MASTER_SPREADSHEET_ID', '')
GOOGLE_CREDENTIALS_JSON = os.environ.get('GOOGLE_CREDENTIALS_JSON', '')

# Catálogo fijo de carriles oficiales.
# carril_values → valores de la columna Carril en la Base Maestra para datasets
#                 con discriminación por sub-carril (Proyectos, Requerimientos).
# org_values    → valores para datasets compartidos sin discriminación (HC, Accesos,
#                 Liberaciones, Riesgos, Dependencias, Replanificaciones, Fases, Estatus).
#                 Cuando no está definido, se usa carril_values.
# req_values    → override específico para Requerimientos cuando difiere de carril_values.
#                 frozenset() vacío = carril sin requerimientos propios.
CARRIL_CATALOG = {
    'mff': {
        'label':           'Money Free Flex',
        'carril_values':   frozenset(['Money Free Flex']),
        'req_cat_values':  frozenset(['Money Free Flex']),
        'hc_aliases':      frozenset(['Money Free Flex']),
    },
    'divisas_bau': {
        'label':           'Divisas BAU',
        'carril_values':   frozenset(['Divisas BAU']),
        'org_values':      frozenset(['Divisas BAU', 'DIVISAS']),
        'req_cat_values':  frozenset(['Divisas']),
        'hc_aliases':      frozenset(['Divisas BAU', 'Divisas Bau']),
    },
    'divisas_transf': {
        'label':           'Divisas Transformación',
        'carril_values':   frozenset(['Divisas Transformación']),
        'req_cat_values':  frozenset(),
        'hc_aliases':      frozenset(['Divisas Transformación', 'Divisas Transforming']),
    },
    'elektra_vales': {
        'label':           'Elektra Vales',
        'carril_values':   frozenset(['Elektra Vales']),
        'req_cat_values':  frozenset(['Tarjeta de Regalo']),
        'hc_aliases':      frozenset(['Elektra Vales / TDR', 'EKTV/TDR']),
    },
    'cripto_corp': {
        'label':           'Cripto Corporativo',
        'carril_values':   frozenset(['Cripto Corporativo']),
        'org_values':      frozenset(['Cripto Corporativo', 'COINPRO']),
        'req_cat_values':  frozenset(),
        'hc_aliases':      frozenset(['Cripto Corporativo', 'Corporativo']),
    },
    'cripto_masivos': {
        'label':           'Cripto Masivos',
        'carril_values':   frozenset(['Cripto Masivos']),
        'org_values':      frozenset(['Cripto Masivos', 'COINPRO']),
        'req_cat_values':  frozenset(['Cripto Masivo']),
        'hc_aliases':      frozenset(['Cripto Masivos', 'Tienda de Monedas']),
    },
    'cripto_backoffice': {
        'label':           'Cripto Backoffice',
        'carril_values':   frozenset(['Cripto Backoffice']),
        'org_values':      frozenset(['Cripto Backoffice', 'COINPRO']),
        'req_cat_values':  frozenset(['Backoffice']),
        'hc_aliases':      frozenset(['Cripto Backoffice', 'Backoffice']),
    },
}

MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
MES_MAP = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',
           7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive',
]


# ── Google Sheets helpers ─────────────────────────────────────────────

def get_workbook():
    if not GOOGLE_CREDENTIALS_JSON:
        raise RuntimeError('GOOGLE_CREDENTIALS_JSON no configurado. Revisa .env o las variables de Render.')
    if not SHEET_ID:
        raise RuntimeError('SHEET_ID no configurado. Revisa .env o las variables de Render.')
    creds = Credentials.from_service_account_info(
        json.loads(GOOGLE_CREDENTIALS_JSON),
        scopes=SCOPES,
    )
    gc = gspread.authorize(creds)
    return gc.open_by_key(SHEET_ID)


def safe_float(v):
    try:
        return float(v) if v not in (None, '', '-') else 0.0
    except (ValueError, TypeError):
        return 0.0


def safe_int(v):
    try:
        return int(float(v)) if v not in (None, '', '-') else 0
    except (ValueError, TypeError):
        return 0


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ── Estáticos ─────────────────────────────────────────────────────────

@app.route('/')
@app.route('/portafolio')
def index():
    resp = send_from_directory(app.static_folder, 'index.html')
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


@app.route('/encuesta')
def encuesta():
    resp = send_from_directory(app.static_folder, 'portal-encuesta-legacy.html')
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp

@app.route('/encuestaAgil')
def encuesta_agil():
    resp = send_from_directory(app.static_folder, 'portal-encuesta.html')
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp


# ── GET /api/encuesta/responses ───────────────────────────────────────

def _norm(name):
    """Normaliza nombre: quita acentos, puntuación, minúsculas para comparar."""
    nfkd = unicodedata.normalize('NFKD', str(name))
    ascii_n = ''.join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r'[^\w\s]', '', ascii_n).lower().strip()

def _edit_dist(a, b):
    """Levenshtein distance entre dos strings (early-exit si > 2)."""
    if abs(len(a) - len(b)) > 2:
        return 99
    m, n = len(a), len(b)
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev, dp[0] = dp[:], i
        for j in range(1, n + 1):
            dp[j] = prev[j-1] if a[i-1] == b[j-1] else 1 + min(prev[j], dp[j-1], prev[j-1])
    return dp[n]

def _fuzzy_match(norm_nombre, padron_lookup):
    """Fallback si no hay match exacto:
    1. Word-subset: todas las palabras del nombre más corto están en el más largo
       (ej. padrón tiene 3 palabras, Sheet tiene 4 → apellido extra).
    2. Edit distance ≤ 1 en string compacto (sin espacios) → typos de 1 letra.
    """
    words_sheet = norm_nombre.split()
    words_sheet_set = set(words_sheet)
    first_sheet = words_sheet[0] if words_sheet else ''
    # 1. Word-subset (requiere ≥2 palabras Y misma primera palabra para evitar falsos positivos)
    if len(words_sheet) >= 2:
        best_key = None
        for pk in padron_lookup:
            pk_words = pk.split()
            pk_set = set(pk_words)
            first_padron = pk_words[0] if pk_words else ''
            if (len(pk_set) >= 2
                    and first_sheet == first_padron
                    and (pk_set <= words_sheet_set or words_sheet_set <= pk_set)):
                if best_key is None or len(pk_words) > len(best_key.split()):
                    best_key = pk
        if best_key:
            return padron_lookup[best_key]
    # 2. Edit distance ≤ 1 en string compacto
    compact = norm_nombre.replace(' ', '')
    for pk, pv in padron_lookup.items():
        if _edit_dist(compact, pk.replace(' ', '')) <= 1:
            return pv
    return None

_PADRON_LOOKUP = None  # nombre_normalizado → (email, udn_html)
_SEED_STATE    = None  # SEED_STATE parseado del HTML
_UDN_DATA      = None  # UDN_DATA parseado del HTML

_GC                  = None   # cliente gspread reutilizable
_ENCUESTA_CACHE      = None   # respuesta cacheada
_ENCUESTA_CACHE_TIME = 0.0    # epoch del último fetch
_ENCUESTA_CACHE_TTL  = 30     # segundos

# Caché por carril — protege contra rate-limit 429 de Google Sheets
_PORTAFOLIO_CACHE          = {}    # carril_key → {'data': dict, 'ts': float}
_PORTAFOLIO_CACHE_TTL      = 600   # 10 minutos — caché principal (stale-while-revalidate)
_PORTAFOLIO_REFRESH_TTL    = 300   # refresco background a los 5 min (sirve datos frescos sin bloquear)
_PORTAFOLIO_REFRESH_LOCK   = set() # carriles con refresco background en curso
_WB_CACHE                  = {}    # sheet_id → Spreadsheet object (evita open_by_key repetido)
_ACCESOS_CACHE             = {'rows': None, 'ts': 0.0}  # caché del tab Accesos (login)
_ACCESOS_CACHE_TTL         = 300   # 5 minutos

def _get_gc():
    global _GC
    if _GC is not None:
        return _GC
    creds = Credentials.from_service_account_info(
        json.loads(GOOGLE_CREDENTIALS_JSON), scopes=SCOPES)
    _GC = gspread.authorize(creds)
    return _GC

def _get_udn_data():
    global _UDN_DATA
    if _UDN_DATA is not None:
        return _UDN_DATA
    html_path = os.path.join(os.path.dirname(__file__), 'static', 'portal-encuesta.html')
    try:
        with open(html_path, encoding='utf-8') as f:
            html = f.read()
        m = re.search(r'const UDN_DATA = ({.*?});\s*\n', html, re.DOTALL)
        _UDN_DATA = json.loads(m.group(1)) if m else {}
    except Exception:
        _UDN_DATA = {}
    return _UDN_DATA

def _get_padron_lookup():
    global _PADRON_LOOKUP
    if _PADRON_LOOKUP is not None:
        return _PADRON_LOOKUP
    udn_data = _get_udn_data()
    lookup = {}
    for udn_key, info in udn_data.items():
        for padron_name, padron_email in info.get('participants', []):
            lookup[_norm(padron_name)] = (padron_email, udn_key)
    _PADRON_LOOKUP = lookup
    return _PADRON_LOOKUP

def _get_seed_state():
    global _SEED_STATE
    if _SEED_STATE is not None:
        return _SEED_STATE
    html_path = os.path.join(os.path.dirname(__file__), 'static', 'portal-encuesta.html')
    try:
        with open(html_path, encoding='utf-8') as f:
            html = f.read()
        m = re.search(r'const SEED_STATE = ({.*?});\s*//', html, re.DOTALL)
        _SEED_STATE = json.loads(m.group(1)) if m else {'responded': [], 'extraParticipants': {}}
    except Exception:
        _SEED_STATE = {'responded': [], 'extraParticipants': {}}
    return _SEED_STATE

# Mapeo de nombres de UDN del Sheet → keys exactos en UDN_DATA del HTML
_SHEET_UDN_MAP = {
    'Satisfacción del Cliente - RGE (SAC)':  'Satisfacción del Cliente - RGE (SAC)',
    'Sistemas operación RGE':                'Satisfacción del Cliente - RGE (SAC)',
    'Seguros Azteca':                        'Seguros',
    'Dirección general de soluciones al cliente, colaborador y cámara de compensación.':
        'Dirección General de Soluciones al Cliente, Colaborador y Cámara de Compensación',
    'BackOffice':                            'Backoffice',
    'Remesas':                               'Remesas',
    'Corporativo (Recaudación, Dispersión, Adquiriente y Corresponsales)':
        'Corporativo (Recaudación, Dispersión, Adquiriente y Corresponsales)',
    'Corporativo (Recaudación, Dispersión, Adquirente)':
        'Corporativo (Recaudación, Dispersión, Adquiriente y Corresponsales)',
    'GS Motos':                              'GS Motos',
    'CPyC - Préstamo':                       'CPyC - Préstamo',
    'Transversales (Negocios de Comisión)':  'Transversales N. Comisión',
    'CPyC - Comercio':                       'CPyC - Comercio',
    'Tesorería':                             'Tesorería',
    'Divisas':                               'Divisas',
    'Money Free Flex':                       'Money Free Flex',
    'Captación / Cuentas de ahorro':         'Captación',
    'BAZ negocio':                           'BAZ negocio',
    'CPyC - Cobranza':                       'CPyC - Cobranza',
    'Información Ejecutiva':                 'Información Ejecutiva',
    'Facturación electrónica':               'Facturación Electrónica',
    'BIG':                                   'BIG',
    'CPyC - Presta Prenda':                  'CPyC - Presta Prenda',
    'Cajeros - ADE':                         'Caja-ADE',
    'Riesgos de crédito':                    'Riesgos de Mercado y Liquidez',
    'Beneficios financieros':                'Beneficios Financieros',
    'Transferencias (SPEI, SPID y TEF)':     'Transferencias (SPEI, SPID y TEF)',
    'Garantía Extendida Elektra':            'Garantía Extendida Elektra',
    'Criptomonedas':                         'Criptomonedas',
    'Afore Azteca':                          'Afore',
    'Círculo de crédito':                    'Círculo de crédito',
    'BAZ entregas':                          'BAZ Entregas',
    'Contraloría':                           'Contraloría Comercio',
    'Franquicia':                            'Bienes Inmuebles - Franquicia',
    'Administración de portafolio':          'Administración de portafolio',
    'Bienestar':                             'Bienestar',
    'Prevención de lavado de dinero':        'Prevención de lavado de dinero',
    'Ecommerce':                             'Ecommerce',
    'Tiempo Aire':                           'Tiempo Aire',
}

@app.route('/api/encuesta/responses')
def encuesta_responses():
    global _GC, _ENCUESTA_CACHE, _ENCUESTA_CACHE_TIME
    sheet_id = os.environ.get('ENCUESTA_SHEET_ID', '')
    if not sheet_id or not GOOGLE_CREDENTIALS_JSON:
        return jsonify({'responded': [], 'extraParticipants': {}, 'total': 0})

    # Devuelve caché si tiene menos de 30 segundos
    now = time.time()
    if _ENCUESTA_CACHE and (now - _ENCUESTA_CACHE_TIME) < _ENCUESTA_CACHE_TTL:
        return jsonify(_ENCUESTA_CACHE)

    try:
        gc     = _get_gc()
        ws     = gc.open_by_key(sheet_id).sheet1
        # Solo pedimos las columnas D:F (correo, nombre, UDN) — 3 cols en vez de 6
        raw    = ws.get('D:F')          # [[header_d, header_e, header_f], [val, val, val], ...]
        padron = _get_padron_lookup()

        if not raw or len(raw) < 2:
            return jsonify({'responded': [], 'extraParticipants': {}, 'total': 0})

        # Detecta posición por nombre de header (resistente a cambios de orden de columnas)
        headers   = raw[0]
        i_email  = next((i for i,h in enumerate(headers) if 'correo'  in h.lower()), 0)
        i_nombre = next((i for i,h in enumerate(headers) if 'nombre'  in h.lower()), 1)
        i_udn    = next((i for i,h in enumerate(headers) if 'unidad'  in h.lower()), 2)

        data_rows          = raw[1:]
        responded          = []
        extra_participants = {}

        for row in data_rows:
            email  = row[i_email ].strip() if i_email  < len(row) else ''
            nombre = row[i_nombre].strip() if i_nombre < len(row) else ''
            udn_s  = row[i_udn   ].strip() if i_udn    < len(row) else ''
            if not email:
                continue

            padron_match = padron.get(_norm(nombre)) or _fuzzy_match(_norm(nombre), padron)
            if padron_match:
                responded.append(padron_match[0])
            else:
                responded.append(email)
                html_udn = _SHEET_UDN_MAP.get(udn_s, udn_s)
                if html_udn:
                    extra_participants.setdefault(html_udn, []).append([nombre, email])

        result = {
            'responded':         responded,
            'extraParticipants': extra_participants,
            'total':             len(data_rows),
        }
        _ENCUESTA_CACHE      = result
        _ENCUESTA_CACHE_TIME = now
        return jsonify(result)
    except Exception as e:
        _GC = None  # fuerza re-auth en el siguiente intento
        app.logger.error(f'[encuesta] Error: {type(e).__name__}: {e}')
        return jsonify({'error': str(e), 'responded': [], 'extraParticipants': {}, 'total': 0}), 500


# ── GET /api/data ─────────────────────────────────────────────────────

@app.route('/api/data')
def get_data():
    """
    Lee Google Sheets y retorna JSON compatible con el dashboard:
      { ultima_actualizacion, developers, proyectos_presupuesto }
    """
    wb = get_workbook()

    # ── Developers ──────────────────────────────────────────────────
    ws_dev    = wb.worksheet('Developers')
    dev_recs  = ws_dev.get_all_records()

    developers = []
    for r in dev_recs:
        nombre = str(r.get('nombre', '') or '').strip()
        if not nombre:
            continue
        developers.append({
            'nombre':      nombre,
            'celula':      str(r.get('celula', '') or '').strip(),
            'tipo':        str(r.get('tipo', 'Sin clasificar') or 'Sin clasificar').strip(),
            'horasPorMes': {m: safe_float(r.get(m, 0)) for m in MESES},
        })

    # ── Proyectos ────────────────────────────────────────────────────
    ws_proy   = wb.worksheet('Proyectos')
    proy_recs = ws_proy.get_all_records()

    proyectos_presupuesto = {}
    for r in proy_recs:
        folio  = str(r.get('folio',  '') or '').strip()
        celula = str(r.get('celula', '') or '').strip()
        if not folio or not celula:
            continue
        proyectos_presupuesto.setdefault(celula, []).append({
            'folio':       folio,
            'nombre':      str(r.get('nombre', '') or '').strip(),
            'esJira':      folio.upper() != 'OTRAS',
            'personasInt': safe_int(r.get('personas_int', 0)),
            'personasExt': safe_int(r.get('personas_ext', 0)),
            'horasPorMes': {m: safe_float(r.get(m, 0)) for m in MESES},
        })

    return jsonify({
        'ultima_actualizacion': datetime.now(timezone.utc).isoformat(),
        'developers':            developers,
        'proyectos_presupuesto': proyectos_presupuesto,
    })


# ── GET /api/health ───────────────────────────────────────────────────

@app.route('/api/health')
def health():
    try:
        wb = get_workbook()
        sheets = [ws.title for ws in wb.worksheets()]
        return jsonify({'status': 'ok', 'hojas': sheets})
    except Exception as e:
        return jsonify({'status': 'error', 'detail': str(e)}), 503


# ── POST /api/import/tempo ────────────────────────────────────────────
# Importa CSV/Excel de Tempo y acumula horas en la hoja Developers.

TEMPO_COLS = {
    'nombre': ['Nombre completo', 'Full name', 'Nombre'],
    'folio':  ['Clave de Incidencia', 'Issue Key', 'Issue key', 'Folio', 'Clave'],
    'horas':  ['Horas', 'Hours', 'Time spent (h)', 'Tiempo invertido (h)'],
    'fecha':  ['Fecha de trabajo', 'Work date', 'Started', 'Fecha'],
}


def find_col(headers, candidates):
    for c in candidates:
        if c in headers:
            return c
    hl = {h.lower(): h for h in headers}
    for c in candidates:
        if c.lower() in hl:
            return hl[c.lower()]
    return None


def parse_csv_rows(b):
    text   = b.decode('utf-8-sig', errors='replace')
    reader = csv.DictReader(io.StringIO(text))
    return list(reader), reader.fieldnames or []


def parse_xlsx_rows(b):
    if not OPENPYXL_OK:
        raise RuntimeError('openpyxl no instalado')
    wb_xl    = openpyxl.load_workbook(io.BytesIO(b), data_only=True)
    ws_xl    = wb_xl.active
    rows_it  = ws_xl.iter_rows(values_only=True)
    headers  = [str(h) if h is not None else '' for h in next(rows_it, [])]
    data     = [{headers[i]: (row[i] if i < len(row) else None)
                 for i in range(len(headers))} for row in rows_it]
    return data, headers


@app.route('/api/import/tempo', methods=['POST'])
def import_tempo():
    if 'file' not in request.files:
        return jsonify({'error': 'No se recibio ningun archivo.'}), 400
    f = request.files['file']
    if not f.filename or not allowed_file(f.filename):
        return jsonify({'error': 'Archivo invalido o extension no permitida.'}), 400

    file_bytes = f.read()
    ext = f.filename.rsplit('.', 1)[1].lower()
    try:
        rows, headers = parse_csv_rows(file_bytes) if ext == 'csv' else parse_xlsx_rows(file_bytes)
    except Exception as e:
        return jsonify({'error': f'No se pudo leer el archivo: {e}'}), 400

    col_nombre = find_col(headers, TEMPO_COLS['nombre'])
    col_horas  = find_col(headers, TEMPO_COLS['horas'])
    col_fecha  = find_col(headers, TEMPO_COLS['fecha'])

    missing = [k for k, c in [('Nombre completo', col_nombre),
                               ('Horas',           col_horas),
                               ('Fecha de trabajo', col_fecha)] if not c]
    if missing:
        return jsonify({'error': f'Columnas no encontradas: {", ".join(missing)}',
                        'detectadas': list(headers)}), 400

    # Acumular horas por (developer, mes)
    acum   = {}
    errores = []
    for i, row in enumerate(rows):
        nombre    = str(row.get(col_nombre, '') or '').strip()
        horas_raw = row.get(col_horas, 0) or 0
        fecha_raw = str(row.get(col_fecha, '') or '')
        if not nombre:
            continue
        try:
            horas = float(str(horas_raw).replace(',', '.'))
        except (ValueError, TypeError):
            errores.append(f'Fila {i+2}: horas invalidas "{horas_raw}"')
            continue
        mes_key = None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y'):
            try:
                mes_key = MES_MAP.get(datetime.strptime(fecha_raw[:10], fmt).month)
                break
            except ValueError:
                continue
        if not mes_key:
            errores.append(f'Fila {i+2}: fecha no reconocida "{fecha_raw}"')
            continue
        acum.setdefault(nombre, {})[mes_key] = acum.get(nombre, {}).get(mes_key, 0.0) + horas

    if not acum:
        return jsonify({'actualizadas': 0, 'errores': errores,
                        'warning': 'No se encontraron filas validas.'})

    wb = get_workbook()
    ws = wb.worksheet('Developers')
    records     = ws.get_all_records()
    header_row  = ws.row_values(1)

    actualizadas   = 0
    no_encontradas = []
    for nombre_dev, meses_data in acum.items():
        row_idx = None
        for idx, r in enumerate(records):
            if str(r.get('nombre', '')).strip().lower() == nombre_dev.lower():
                row_idx = idx + 2
                break
        if row_idx is None:
            no_encontradas.append(nombre_dev)
            continue
        for mes, horas in meses_data.items():
            if mes in header_row:
                col_idx = header_row.index(mes) + 1
                current = safe_float(ws.cell(row_idx, col_idx).value)
                ws.update_cell(row_idx, col_idx, round(current + horas, 2))
                actualizadas += 1

    return jsonify({
        'actualizadas':  actualizadas,
        'errores':       errores,
        'no_encontradas': no_encontradas,
    })


# ── Portafolio: helpers ───────────────────────────────────────────────

def get_master_workbook():
    if not GOOGLE_CREDENTIALS_JSON:
        raise RuntimeError('GOOGLE_CREDENTIALS_JSON no configurado.')
    if not MASTER_SPREADSHEET_ID:
        raise RuntimeError('MASTER_SPREADSHEET_ID no configurado. Agrega la var en Render.')
    if MASTER_SPREADSHEET_ID not in _WB_CACHE:
        _WB_CACHE[MASTER_SPREADSHEET_ID] = _get_gc().open_by_key(MASTER_SPREADSHEET_ID)
    return _WB_CACHE[MASTER_SPREADSHEET_ID]


def sem_to_label(s):
    """'Sem 27-2026' → 'SEM 27 · 2026'. '28' → 'SEM 28 · 2026'. '-' → '—'."""
    s = str(s or '').strip()
    if not s or s.lower() in ('-', '—', 'tbd', 'n/a', 'na', '-'):
        return '—'
    m = re.match(r'[Ss]em\s*(\d+)-(\d{4})', s)
    if m:
        return f'SEM {m.group(1)} · {m.group(2)}'
    if re.match(r'^\d{1,2}$', s):
        return f'SEM {s} · 2026'
    return s.upper()


def col_index_to_a1(i):
    """0-based column index → A1 column letter(s)."""
    s = ''
    idx = i + 1
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(ord('A') + r) + s
    return s


def week_to_date(week_num, year=2026, end_of_week=False):
    """ISO week number → YYYY-MM-DD (Monday by default, Friday when end_of_week=True)."""
    try:
        w = int(week_num)
        if w < 1:
            return None
    except (ValueError, TypeError):
        return None
    jan4 = date(year, 1, 4)
    monday_w1 = jan4 - timedelta(days=jan4.weekday())
    result = monday_w1 + timedelta(weeks=w - 1)
    if end_of_week:
        result += timedelta(days=4)  # Monday → Friday
    return result.isoformat()


COLOR_MAP = {
    'amarillo': '#f59e0b',
    'verde':    '#16a34a',
    'rojo':     '#dc2626',
    'azul':     '#1d4ed8',
    'morado':   '#7c3aed',
    'naranja':  '#d97706',
    'gris':     '#6b7280',
    'lila':     '#7f77dd',
    'cyan':     '#0ea5e9',
}

# Normalización de Estatus_Actividad → vocabulario del Dashboard
_ESTATUS_NORM = {
    'en curso':    'En progreso',
    'en progreso': 'En progreso',
    'finalizado':  'Hecho',
    'hecho':       'Hecho',
    'por hacer':   'Por hacer',
    'por_hacer':   'Por hacer',
}

def norm_estatus_actividad(s):
    """Normaliza Estatus_Actividad al vocabulario exacto del Dashboard."""
    if not s:
        return None
    normalized = _ESTATUS_NORM.get(s.lower().strip())
    return normalized if normalized else (s.strip() or None)

SEMAFORO_COLOR = {
    'verde':    '#16a34a',
    'amarillo': '#d97706',
    'rojo':     '#dc2626',
}

ROLE_COLORS = {
    'negocio':        {'bg': '#dbeafe', 'fg': '#1d4ed8'},
    'sistemas':       {'bg': '#d1fae5', 'fg': '#065f46'},
    'implementación': {'bg': '#fce7f3', 'fg': '#9d174d'},
    'implementacion': {'bg': '#fce7f3', 'fg': '#9d174d'},
    'agilidad':       {'bg': '#ede9fe', 'fg': '#5b21b6'},
    'scrum master':   {'bg': '#ede9fe', 'fg': '#5b21b6'},
    'business analyst': {'bg': '#fef9c3', 'fg': '#854d0e'},
}


def color_to_hex(s):
    return COLOR_MAP.get(str(s or '').lower().strip(), '#6b7280')


def make_resp(rol, nombre):
    c = ROLE_COLORS.get(rol.lower().strip(), {'bg': '#f3f4f6', 'fg': '#374151'})
    words = (nombre or '').strip().split()
    av = ((words[0][0] if words else '') + (words[1][0] if len(words) > 1 else '')).upper()
    return {'rol': rol, 'nombre': nombre, 'av': av, 'bg': c['bg'], 'fg': c['fg']}


def sheet_rows(worksheet, head_row=1):
    """Return list of dicts using the given 1-indexed header row."""
    all_vals = worksheet.get_all_values()
    if len(all_vals) < head_row:
        return []
    headers = all_vals[head_row - 1]
    return [
        {headers[i]: (row[i] if i < len(row) else '')
         for i in range(len(headers))}
        for row in all_vals[head_row:]
    ]


def _fetch_tabs_batch(wb, tab_specs):
    """Lee varios tabs del workbook en UNA sola llamada HTTP (batchGet).
    tab_specs: [(nombre, head_row), ...]
    Devuelve: {nombre: [dict, ...]}  — tabs ausentes devuelven [].
    """
    ranges = ["'{}'!A:ZZ".format(name) for name, _ in tab_specs]
    try:
        resp = wb.values_batch_get(ranges, params={'valueRenderOption': 'FORMATTED_VALUE'})
        value_ranges = resp.get('valueRanges', [])
    except Exception as e:
        app.logger.warning(f'[portafolio] batch_get falló ({e}); leyendo tabs individualmente')
        result = {}
        for name, hr in tab_specs:
            try:
                result[name] = sheet_rows(wb.worksheet(name), hr)
            except (gspread.exceptions.WorksheetNotFound, gspread.exceptions.APIError):
                result[name] = []
        return result

    result = {}
    for (name, head_row), vr in zip(tab_specs, value_ranges):
        all_vals = vr.get('values', [])
        if len(all_vals) < head_row:
            result[name] = []
            continue
        headers = all_vals[head_row - 1]
        result[name] = [
            {headers[i]: (row[i] if i < len(row) else '')
             for i in range(len(headers))}
            for row in all_vals[head_row:]
        ]
    return result


def _filter_by_carril(rows: list, carril_values: frozenset) -> list:
    if not carril_values:
        return []
    return [r for r in rows if r.get('Carril', '') in carril_values]


def _filter_hc_by_celula(rows: list, hc_aliases: frozenset) -> list:
    if not hc_aliases:
        return []
    return [r for r in rows if str(r.get('CÉLULA.', '') or '').strip() in hc_aliases]


def parse_sem_weeks(sem_str):
    """'27-28' or '27' → list of ints. '-' or '' → []"""
    s = str(sem_str or '').strip()
    if not s or s == '-':
        return []
    if '-' in s:
        parts = s.split('-')
        try:
            return list(range(int(parts[0]), int(parts[1]) + 1))
        except (ValueError, IndexError):
            pass
    try:
        return [int(s)]
    except ValueError:
        return []


# ── GET /api/portafolio/data ──────────────────────────────────────────

_EMPTY_CARRIL_RESPONSE = {
    'carriles': {}, 'proyectos': {}, 'portafolios': {},
    'liberaciones': [], 'hc_roster': [],
}

def _portafolio_refresh_bg(carril_key, cfg):
    """Hilo background: re-lee la Base Maestra y actualiza caché sin bloquear requests."""
    try:
        wb   = get_master_workbook()
        data = _build_portafolio_data(wb, cfg, carril_key)
        _PORTAFOLIO_CACHE[carril_key] = {'data': data, 'ts': time.time()}
        app.logger.info(f'[portafolio] Refresco background completado: {carril_key}')
    except Exception as e:
        app.logger.error(f'[portafolio] Error en refresco background {carril_key}: {e}')
    finally:
        _PORTAFOLIO_REFRESH_LOCK.discard(carril_key)


def _build_portafolio_data(wb, cfg, carril_key):
    """Lee la Base Maestra y construye el dict de respuesta filtrando por columna Carril."""
    # ── Lectura batch: tabs presentes en la Base Maestra ─────────────
    # Tabs actuales (9): Accesos, Head Count, Proyectos, Responsables, Riesgos,
    # Dependencias, Replanificaciones, Requerimientos, Liberaciones.
    # 'Estatus' y 'Fases' ya no existen en la BM — se eliminaron del batch.
    _tabs = _fetch_tabs_batch(wb, [
        ('Proyectos', 1), ('Riesgos', 1),
        ('Dependencias', 1), ('Replanificaciones', 1),
        ('Requerimientos', 1), ('Liberaciones', 1), ('Head Count', 1),
    ])

    # ── Conjuntos de filtrado por tipo de dataset ─────────────────────
    # carril_vals:    datasets con columna Carril (Proyectos, Riesgos, etc.)
    # org_vals:       datasets con columna Carril al nivel de sistema origen.
    # req_cat_vals:   Requerimientos — columna Categoria_Etiqueta (NO Carril).
    carril_vals    = cfg['carril_values']
    org_vals       = cfg.get('org_values', carril_vals)
    req_cat_vals   = cfg.get('req_cat_values', frozenset())

    rows_proy = _filter_by_carril(_tabs.get('Proyectos',         []), carril_vals)
    rows_rie  = _filter_by_carril(_tabs.get('Riesgos',           []), org_vals)
    rows_dep  = _filter_by_carril(_tabs.get('Dependencias',      []), org_vals)
    rows_rep  = _filter_by_carril(_tabs.get('Replanificaciones', []), org_vals)
    # Requerimientos: el carril del requerimiento se identifica por Categoria_Etiqueta
    # (la pestaña Requerimientos no tiene columna Carril). La comparación se hace de
    # forma robusta (sin acentos/mayúsculas/espacios) contra la IDENTIDAD del carril
    # —req_cat_values ∪ carril_values ∪ label— para NO depender de una lista manual
    # (req_cat_values) que puede quedar incompleta. Ésa era la causa raíz por la que
    # carriles como "Cripto Corporativo" (con req_cat_values vacío) aparecían sin
    # requerimientos aunque sí existían en el Sheet.
    _req_match = {_norm(v) for v in (set(req_cat_vals) | set(carril_vals) | {cfg.get('label', '')}) if v}
    _all_req   = _tabs.get('Requerimientos', [])
    rows_req   = [r for r in _all_req if _norm(r.get('Categoria_Etiqueta', '')) in _req_match]
    rows_lib  = _filter_by_carril(_tabs.get('Liberaciones',      []), org_vals)
    hc_aliases = cfg.get('hc_aliases', frozenset())
    rows_hc    = _filter_hc_by_celula(_tabs.get('Head Count',   []), hc_aliases)

    # Estatus_Manual: leer del tab Estatus (creado por PATCH; puede no existir aún)
    estatus_manual_map = {}
    try:
        ws_est    = wb.worksheet('Estatus')
        rows_est  = sheet_rows(ws_est, 1)
        for r in rows_est:
            f = str(r.get('ID_Proyecto', '') or '').strip()
            a = str(r.get('Actividad',   '') or '').strip()
            m = str(r.get('Estatus',     '') or '').strip()
            if f and a:
                estatus_manual_map[(f, a)] = m
    except gspread.exceptions.WorksheetNotFound:
        pass
    estatus_override = {}

    # ── 1. PROYECTOS (Gantt) ─────────────────────────────────────────

    proyectos  = {}   # folio → proyecto
    proy_order = []   # insertion order

    for r in rows_proy:
        folio = str(r.get('ID_Proyecto', '') or '').strip()
        if not folio:
            continue

        if folio not in proyectos:
            proy_order.append(folio)
            resp = []
            for rol, key in [
                ('Negocio',        'Responsable_Negocio_Nombre'),
                ('Sistemas',       'Responsable_Sistemas_Nombre'),
                ('Implementación', 'Responsable_Implementacion_Nombre'),
                ('Agilidad',       'Responsable_Agilidad_Nombre'),
            ]:
                nombre = str(r.get(key, '') or '').strip()
                if nombre:
                    resp.append(make_resp(rol, nombre))

            semaforo_raw = str(r.get('Estado_Semaforo', '') or '').strip() or None
            proyectos[folio] = {
                'folio':     folio,
                'nombre':    str(r.get('Nombre_Proyecto', '') or '').strip() or None,
                'showDates': True,
                'carril':    str(r.get('Carril', '') or '').strip() or None,
                'estado':    str(r.get('Estado_Descripcion', '') or '').strip() or None,
                'semaforo':  semaforo_raw,
                'trimestre': str(r.get('Trimestre', '') or '').strip() or None,
                'negocio':   str(r.get('Categoria_Etiqueta', '') or '').strip() or None,
                'semIni':    sem_to_label(r.get('Fecha_Inicio_Proyecto', '')),
                'semFin':    sem_to_label(r.get('Fecha_Fin_Proyecto', '')),
                'semReplan': sem_to_label(r.get('Fecha_Replanificada', '')),
                'resp':      resp,
                'hcI':       0,
                'hcE':       0,
                'numR':      safe_int(r.get('Num_Replanificaciones', 0)),
                'motivoR':   str(r.get('Motivo_Replanificacion', '') or '').strip() or None,
                'kpi':       {'r': 0, 'rA': 0, 'rM': 0, 'd': 0, 'dI': 0, 'dE': 0},
                '_grupos_map': {},
                '_grupos_order': [],
                'replan':      [],
                'riesgos':     [],
                'deps':        [],
                'fase_manual': None,
            }

        # Activity row
        act_name   = str(r.get('Actividad', '') or '').strip()
        grupo_lbl  = str(r.get('Etiqueta_Barra', '') or '').strip()
        color_raw  = str(r.get('Color_Estado_Barra', '') or '').strip()
        sem_ini_a  = str(r.get('Semana_Inicio_Actividad', '') or '').strip()
        sem_fin_a  = str(r.get('Semana_Fin_Actividad', '') or '').strip()

        # Si Actividad está vacía pero Etiqueta_Barra está llena → usar como nombre de la actividad.
        # Esto ocurre cuando el equipo deja la columna Actividad en blanco y solo llena la barra.
        if not act_name and grupo_lbl:
            act_name = grupo_lbl

        if not act_name or not grupo_lbl:
            continue

        year_match = re.search(r'(\d{4})', str(r.get('Fecha_Inicio_Proyecto', '') or ''))
        year = int(year_match.group(1)) if year_match else 2026

        proy = proyectos[folio]
        gmap = proy['_grupos_map']
        if grupo_lbl not in gmap:
            proy['_grupos_order'].append(grupo_lbl)
            gmap[grupo_lbl] = {'label': grupo_lbl, 'color': color_to_hex(color_raw), 'acts': []}

        ini_date    = week_to_date(sem_ini_a, year)
        fin_date    = week_to_date(sem_fin_a, year, end_of_week=True)
        # Estatus: tab "Estatus" (override del usuario) > columna Estatus_Actividad normalizada
        raw_estatus = str(r.get('Estatus_Actividad', '') or '').strip() or None
        estatus_val = estatus_override.get((folio, act_name)) or norm_estatus_actividad(raw_estatus)
        # Fase: valor exacto de la columna "Fase del proyecto" (no se infiere)
        fase_val    = str(r.get('Fase del proyecto', '') or '').strip() or None
        if ini_date:
            gmap[grupo_lbl]['acts'].append({
                'name':           act_name,
                'ini':            ini_date,
                'fin':            fin_date or ini_date,
                'color':          color_to_hex(color_raw),
                'estatus':        estatus_val,
                'estatus_manual': estatus_manual_map.get((folio, act_name), ''),
                'fase':           fase_val,
            })

    # Finalise grupos
    for folio, proy in proyectos.items():
        proy['grupos'] = [proy['_grupos_map'][l] for l in proy['_grupos_order']]
        del proy['_grupos_map']
        del proy['_grupos_order']

    # CARRILES: carril → [folios]
    carriles = {}
    for folio in proy_order:
        carril = proyectos[folio]['carril']
        carriles.setdefault(carril, [])
        if folio not in carriles[carril]:
            carriles[carril].append(folio)

    # ── 1b. RIESGOS ──────────────────────────────────────────────────
    if rows_rie:
        for r in rows_rie:
            folio = str(r.get('ID_Proyecto', '') or '').strip()
            titulo = str(r.get('Titulo', '') or '').strip()
            if not folio or not titulo or folio not in proyectos:
                continue
            nivel = str(r.get('Nivel', '') or '').strip()
            proyectos[folio]['riesgos'].append({
                'titulo':   titulo,
                'nivel':    nivel,
                'fechaImp': str(r.get('Fecha_Impacto', '') or '').strip(),
                'impacto':  str(r.get('Impacto', '') or '').strip(),
                'plan':     str(r.get('Plan_Mitigacion', '') or '').strip(),
                'resp':     str(r.get('Responsable', '') or '').strip(),
            })
            k = proyectos[folio]['kpi']
            k['r'] += 1
            if nivel.lower() == 'alto':
                k['rA'] += 1
            else:
                k['rM'] += 1
    # ── 1c. DEPENDENCIAS ─────────────────────────────────────────────
    if rows_dep:
        for r in rows_dep:
            folio = str(r.get('ID_Proyecto', '') or '').strip()
            titulo = str(r.get('Titulo', '') or '').strip()
            if not folio or not titulo or folio not in proyectos:
                continue
            tipo = str(r.get('Tipo', '') or '').strip()
            proyectos[folio]['deps'].append({
                'titulo': titulo,
                'tipo':   tipo,
                'fecha':  str(r.get('Fecha_Compromiso', '') or '').strip(),
                'desc':   str(r.get('Descripcion', '') or '').strip(),
                'resp':   str(r.get('Responsable', '') or '').strip(),
            })
            k = proyectos[folio]['kpi']
            k['d'] += 1
            if tipo.lower() == 'interna':
                k['dI'] += 1
            else:
                k['dE'] += 1
    # ── 1d. REPLANIFICACIONES ─────────────────────────────────────────
    rep_by_folio = {}
    _FECHA_SKIP = {'n/a', 'na', '-', '—', 'sin registro', 'ninguna'}
    for r in rows_rep:
        folio = str(r.get('ID_Proyecto', '') or '').strip()
        fecha = str(r.get('Fecha', '') or '').strip()
        if not folio or not fecha or fecha.lower() in _FECHA_SKIP or folio not in proyectos:
            continue
        rep_by_folio.setdefault(folio, []).append({
            'fecha':       fecha,
            'motivo':      str(r.get('Motivo', '') or '').strip(),
            'responsable': str(r.get('Responsable', '') or '').strip(),
            'semP':        str(r.get('Semana_Anterior', '') or '').strip(),
            'semN':        str(r.get('Semana_Nueva', '') or '').strip(),
            'reciente':    False,
        })
    # Sync numR from tab count (tab is authoritative; ignore Proyectos sheet column)
    for folio in proyectos:
        items = rep_by_folio.get(folio, [])
        if items:
            items[-1]['reciente'] = True
        proyectos[folio]['replan']  = items
        proyectos[folio]['numR']    = len(items)
        proyectos[folio]['motivoR'] = items[-1]['motivo'] if items else 'Sin replanificaciones.'

    # ── 1e. FASES MANUALES ───────────────────────────────────────────
    # Fase_Actual: override manual del semáforo, leído del tab 'Fases'
    # (creado por PATCH /api/portafolio/proyecto/fase; puede no existir aún).
    try:
        ws_fase   = wb.worksheet('Fases')
        rows_fase = sheet_rows(ws_fase, 1)
        for r in rows_fase:
            f = str(r.get('ID_Proyecto', '') or '').strip()
            fa = str(r.get('Fase_Actual', '') or '').strip()
            if f and fa and f in proyectos:
                proyectos[f]['fase_manual'] = fa
    except gspread.exceptions.WorksheetNotFound:
        pass

    # ── 2. REQUERIMIENTOS → PORTAFOLIOS ─────────────────────────────
    portafolios = {}
    # Carril key = cfg['label'] (display name del carril); todos los items del
    # filtrado req_cat_vals pertenecen al mismo carril, Categoria_Etiqueta es solo
    # el discriminador de filtrado — no un sub-agrupador.
    carril_label = cfg['label']
    _REQ_ESTADO_NORM = {'pleneado': 'Planeado', 'en firmas': 'En firmas'}

    for i, r in enumerate(rows_req):
        folio  = str(r.get('ID_Proyecto', '') or '').strip()
        nombre = str(r.get('Nombre_Proyecto', '') or '').strip()
        # Saltar filas decorativas o sin folio real
        if not folio or folio.lower() == 'sin folio':
            continue
        if not any(c.isdigit() or c == '-' for c in folio):
            continue

        pf = portafolios.setdefault(carril_label, {
            'id':        carril_label.lower().replace(' ', '_'),
            'nombre':    carril_label,
            'trimestre': str(r.get('Trimestre', 'Q3') or 'Q3').strip() + '-2026',
            'items':     [],
        })

        prio_raw = str(r.get('Prioridad', '') or '').strip()
        try:
            prioridad = int(prio_raw) if prio_raw and prio_raw.isdigit() else None
        except ValueError:
            prioridad = None

        estado_raw = str(r.get('ESTADO', '') or r.get('Estado', '') or '').strip()
        estado_norm = _REQ_ESTADO_NORM.get(estado_raw.lower(), estado_raw)

        pf['items'].append({
            'id':               i + 1,
            'prioridad':        prioridad,
            'folio':            folio,
            'nombre':           nombre,
            'descripcion':      str(r.get('Descripción', '') or r.get('Descripcion', '') or '').strip(),
            'solicitante':      str(r.get('Solicitante', '') or '').strip(),
            'ba':               str(r.get('Bussines Analyst', '') or '').strip() or None,
            'num_refinamientos':safe_int(r.get('Números de refinamientos', 0)),
            'estado':           estado_norm,
            'deadline':         str(r.get('Deadline', '') or '').strip() or None,
            'ots':              [],
        })

    # ── 3. LIBERACIONES ─────────────────────────────────────────────
    trenes = {}
    tren_order = []

    for r in rows_lib:
        tid = str(r.get('ID_Tren', '') or '').strip()
        if not tid:
            continue
        if tid not in trenes:
            tren_order.append(tid)
            trenes[tid] = {
                'id':        tid,
                'name':      str(r.get('Nombre_Tren', '') or tid).strip(),
                'ann_sem':   safe_int(r.get('Sem_Anuncio', 0)),
                'qa_str':    str(r.get('Sem_QA', '') or '').strip(),
                'beta_str':  str(r.get('Sem_BETA', '') or '').strip(),
                'tda_sem':   safe_int(r.get('SemTienda', 0)),
                'estado':    str(r.get('Estado_Tren', '') or '').strip(),
                '_cambios':  {},
                '_c_order':  [],
            }
        t = trenes[tid]
        id_cambio   = str(r.get('ID_Cambio',          '') or '').strip()
        cambio      = str(r.get('Cambio',              '') or '').strip()
        tipo        = str(r.get('Tipo',                '') or '').strip()
        folio_raw   = str(r.get('Folio',               '') or '').strip()
        hito        = str(r.get('Hito',                '') or '').strip()
        dependencia = str(r.get('Dependencia',         '') or '').strip()
        estatus_dep = str(r.get('Estatus dependencia', '') or '').strip()
        if not id_cambio:
            continue
        folio = folio_raw if folio_raw not in ('', '-') else ''
        def _clean(v): return v if v not in ('', '-') else ''
        item_key = id_cambio
        if item_key not in t['_cambios']:
            t['_c_order'].append(item_key)
            t['_cambios'][item_key] = {
                'id_cambio':   id_cambio,
                'cambio':      cambio,
                'folio':       folio,
                'hito':        _clean(hito),
                'dependencia': _clean(dependencia),
                'estatus_dep': _clean(estatus_dep),
                'tracks':      [],
            }
        if tipo and tipo not in [tr['label'] for tr in t['_cambios'][item_key]['tracks']]:
            t['_cambios'][item_key]['tracks'].append({'label': tipo, 'milestones': {}})

    liberaciones = []
    for tid in tren_order:
        t = trenes[tid]
        ann = t['ann_sem']
        tda = t['tda_sem']
        qa_weeks   = parse_sem_weeks(t['qa_str'])
        beta_weeks = parse_sem_weeks(t['beta_str'])
        all_nums = sorted(set(
            ([ann] if ann else []) + qa_weeks + beta_weeks + ([tda] if tda else [])
        ))
        if not all_nums:
            continue
        weeks = [{'num': w} for w in all_nums]
        idx_map = {w: i for i, w in enumerate(all_nums)}

        items = []
        for cid in t['_c_order']:
            cb = t['_cambios'][cid]
            tracks = []
            for tr in cb['tracks']:
                ms = {}
                if ann in idx_map:
                    ms[idx_map[ann]] = 'anunciado'
                for w in qa_weeks:
                    if w in idx_map:
                        ms[idx_map[w]] = 'qa'
                for w in beta_weeks:
                    if w in idx_map:
                        ms[idx_map[w]] = 'beta'
                if tda and tda in idx_map:
                    ms[idx_map[tda]] = 'tienda'
                # Convert int keys to strings for JSON compatibility
                tracks.append({'label': tr['label'], 'milestones': {str(k): v for k, v in ms.items()}})
            items.append({
                'id_cambio':   cb['id_cambio'],
                'folio':       cb['folio'],
                'cambio':      cb['cambio'],
                'hito':        cb['hito'],
                'dependencia': cb['dependencia'],
                'estatus_dep': cb['estatus_dep'],
                'tracks':      tracks,
            })

        liberaciones.append({
            'id':           tid,
            'name':         t['name'],
            'estado':       t['estado'],
            'anuncioSem':   ann,
            'liberacionSem': tda or ann,
            'weeks':        weeks,
            'items':        items,
        })

    # ── 4. HEAD COUNT ────────────────────────────────────────────────
    _MES_ES = {'ene':1,'feb':2,'mar':3,'abr':4,'may':5,'jun':6,
               'jul':7,'ago':8,'sep':9,'oct':10,'nov':11,'dic':12}

    def parse_hc_date(s):
        """Parsea fechas en varios formatos a YYYY-MM-DD.
        Google Sheets devuelve fechas en formato US (MM/DD/YYYY) via get_all_values().
        Si el primer componente > 12, forzosamente es el día (DD/MM); de lo contrario MM/DD."""
        if not s:
            return None
        s = str(s).strip()
        # YYYY-MM-DD
        if re.match(r'^\d{4}-\d{2}-\d{2}', s):
            return s[:10]
        # X/Y/YYYY — detectar formato: si X>12 forzosamente es DD/MM; si no, asumir DD/MM (formato MX)
        m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
        if m:
            a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if a > 12:   # primer componente no puede ser mes → DD/MM
                mo, d2 = b, a
            else:        # asumir DD/MM (formato estándar MX)
                mo, d2 = b, a
            return f'{y}-{mo:02d}-{d2:02d}'
        # X/Y/YY (año de 2 dígitos) — misma lógica
        m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{2})$', s)
        if m:
            a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            full_y = 2000 + y if y < 50 else 1900 + y
            if a > 12:
                mo, d2 = b, a
            else:        # asumir DD/MM (formato estándar MX)
                mo, d2 = b, a
            return f'{full_y}-{mo:02d}-{d2:02d}'
        # DD-MMM-YY o DD-MMM-YYYY (meses en español: abr, ago, sep…)
        m = re.match(r'^(\d{1,2})[-/]([a-zA-Z]{3})[-/](\d{2,4})$', s)
        if m:
            d2, mes, y = int(m.group(1)), m.group(2).lower(), int(m.group(3))
            mo = _MES_ES.get(mes)
            if mo:
                full_y = 2000 + y if y < 100 else y
                return f'{full_y}-{mo:02d}-{d2:02d}'
        return None

    def extract_folio(proceso_str):
        """Extraer folio LEANBA-XXXXXX del campo PROYECTO/PROCESO."""
        if not proceso_str:
            return None
        m = re.match(r'(LEANBA[-]?\d+)', proceso_str, re.IGNORECASE)
        return m.group(1) if m else None

    hc_roster = []
    for r in rows_hc:
        nombre = str(r.get('NOMBRE DESARROLLADOR.', '') or '').strip()
        if not nombre:
            continue
        celula  = str(r.get('CÉLULA.', '') or '').strip()
        negocio = str(r.get('NEGOCIO.', '') or '').strip().rstrip('.')
        rol     = str(r.get('ROL.', '') or '').strip()
        esquema = str(r.get('INTERNO/EXTERNO', '') or '').strip()
        perfil  = str(r.get('Perfil', '') or r.get('PERFIL', '') or '').strip()
        proceso      = str(r.get('PROYECTO/PROCESO', '') or '').strip()
        folio_direct = str(r.get('FOLIO', '') or '').strip()
        f_ini        = parse_hc_date(str(r.get('FECHA INICIO', '') or '').strip())
        f_fin        = parse_hc_date(str(r.get('FECHA FIN', '') or '').strip())
        folio        = folio_direct or extract_folio(proceso)

        proyectos_hc = []
        if proceso or folio:
            proyectos_hc = [{'folio': folio, 'desc': proceso, 'ini': f_ini, 'fin': f_fin}]

        hc_roster.append({
            'nombre':    nombre,
            'carril':    celula,
            'negocio':   negocio,
            'rol':       rol,
            'esquema':   esquema,
            'perfil':    perfil,
            'proyectos': proyectos_hc,
        })

    # Calcular HC por proyecto desde el roster
    for folio_p, proy in proyectos.items():
        folio_norm = re.sub(r'[-]', '', folio_p).upper()
        hci = sum(1 for d in hc_roster
                  if d['rol'] != 'Gerente' and d['esquema'] == 'Interno'
                  and any(re.sub(r'[-]', '', (pr.get('folio') or '')).upper() == folio_norm
                          for pr in d['proyectos']))
        hce = sum(1 for d in hc_roster
                  if d['rol'] != 'Gerente' and d['esquema'] == 'Externo'
                  and any(re.sub(r'[-]', '', (pr.get('folio') or '')).upper() == folio_norm
                          for pr in d['proyectos']))
        if hci or hce:
            proy['hcI'] = hci
            proy['hcE'] = hce

    return {
        'ultima_actualizacion': datetime.now(timezone.utc).isoformat(),
        'carriles':      carriles,
        'proyectos':     proyectos,
        'portafolios':   portafolios,
        'liberaciones':  liberaciones,
        'hc_roster':     hc_roster,
    }


# ── GET /api/portafolio/data ─────────────────────────────────────────────────

@app.route('/api/portafolio/data')
def get_portafolio_data():
    carril_key = request.args.get('carril', 'mff').strip().lower()
    cfg        = CARRIL_CATALOG.get(carril_key)
    now        = time.time()

    if cfg is None:
        return jsonify({'ultima_actualizacion': datetime.now(timezone.utc).isoformat(),
                        **_EMPTY_CARRIL_RESPONSE})

    cached = _PORTAFOLIO_CACHE.get(carril_key)
    age    = now - cached['ts'] if cached else None

    # 1. Caché fresco → respuesta inmediata
    if cached and age < _PORTAFOLIO_REFRESH_TTL:
        return jsonify(cached['data'])

    # 2. Caché algo viejo pero válido → servir inmediatamente + refresco en background
    if cached and age < _PORTAFOLIO_CACHE_TTL:
        if carril_key not in _PORTAFOLIO_REFRESH_LOCK:
            _PORTAFOLIO_REFRESH_LOCK.add(carril_key)
            threading.Thread(
                target=_portafolio_refresh_bg,
                args=(carril_key, cfg),
                daemon=True,
            ).start()
        return jsonify(cached['data'])

    # 3. Sin caché o expirado → lectura síncrona (sólo al arrancar por primera vez)
    try:
        wb = get_master_workbook()
    except Exception as e:
        app.logger.error(f'[portafolio] Error abriendo Base Maestra para {carril_key}: {e}')
        if cached:
            return jsonify(cached['data'])
        return jsonify({'ultima_actualizacion': datetime.now(timezone.utc).isoformat(),
                        **_EMPTY_CARRIL_RESPONSE})

    try:
        data = _build_portafolio_data(wb, cfg, carril_key)
    except Exception as e:
        app.logger.error(f'[portafolio] Error construyendo datos {carril_key}: {e}')
        if cached:
            return jsonify(cached['data'])
        return jsonify({'ultima_actualizacion': datetime.now(timezone.utc).isoformat(),
                        **_EMPTY_CARRIL_RESPONSE})

    _PORTAFOLIO_CACHE[carril_key] = {'data': data, 'ts': time.time()}
    return jsonify(data)


# ── PATCH /api/portafolio/actividad/estatus ───────────────────────────

@app.route('/api/portafolio/actividad/estatus', methods=['PATCH'])
def patch_actividad_estatus():
    data      = request.get_json(silent=True) or {}
    folio     = str(data.get('folio',     '') or '').strip()
    actividad = str(data.get('actividad', '') or '').strip()
    estatus   = str(data.get('estatus',   '') or '').strip()
    emp       = str(data.get('emp',       '') or '').strip()

    VALID_ESTATUS = {'Por hacer', 'En progreso', 'Hecho', ''}
    if not folio or not actividad or estatus not in VALID_ESTATUS:
        return jsonify({'error': 'Datos inválidos'}), 400

    wb = get_master_workbook()
    try:
        ws = wb.worksheet('Estatus')
    except gspread.exceptions.WorksheetNotFound:
        ws = wb.add_worksheet(title='Estatus', rows=500, cols=5)
        ws.append_row(['ID_Proyecto', 'Actividad', 'Estatus', 'Actualizado_Em', 'Fecha_Actualizacion'],
                      value_input_option='RAW')

    all_vals = ws.get_all_values()
    headers  = all_vals[0] if all_vals else []
    col_map  = {h: i for i, h in enumerate(headers)}
    ci_folio = col_map.get('ID_Proyecto', 0)
    ci_act   = col_map.get('Actividad',   1)
    ci_est   = col_map.get('Estatus',     2)
    ci_emp   = col_map.get('Actualizado_Em', 3)
    ci_fecha = col_map.get('Fecha_Actualizacion', 4)

    now_str = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')

    for row_i, row in enumerate(all_vals[1:], start=2):
        r_folio = row[ci_folio].strip() if len(row) > ci_folio else ''
        r_act   = row[ci_act].strip()   if len(row) > ci_act   else ''
        if r_folio == folio and r_act == actividad:
            ws.update_acell(col_index_to_a1(ci_est)   + str(row_i), estatus)
            ws.update_acell(col_index_to_a1(ci_emp)   + str(row_i), emp)
            ws.update_acell(col_index_to_a1(ci_fecha) + str(row_i), now_str)
            return jsonify({'ok': True, 'updated': True})

    ws.append_row([folio, actividad, estatus, emp, now_str], value_input_option='USER_ENTERED')
    return jsonify({'ok': True, 'created': True})


# ── POST /api/auth/login ──────────────────────────────────────────────

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    data = request.get_json(silent=True) or {}
    emp  = str(data.get('emp',  '') or '').strip()
    pw   = str(data.get('pass', '') or '').strip()

    if not emp or not pw:
        return jsonify({'ok': False, 'error': 'Credenciales vacías'}), 400

    # Usar caché del tab Accesos para no bloquear en cada intento de login
    now = time.time()
    if _ACCESOS_CACHE['rows'] is None or (now - _ACCESOS_CACHE['ts']) > _ACCESOS_CACHE_TTL:
        try:
            wb   = get_master_workbook()
            ws   = wb.worksheet('Accesos')
            _ACCESOS_CACHE['rows'] = sheet_rows(ws, head_row=1)
            _ACCESOS_CACHE['ts']   = now
        except gspread.exceptions.WorksheetNotFound:
            return jsonify({'ok': False, 'error': 'Tab Accesos no encontrado'}), 500
        except Exception as e:
            app.logger.error(f'[auth] Error leyendo Accesos: {e}')
            if _ACCESOS_CACHE['rows'] is None:
                return jsonify({'ok': False, 'error': 'Error de autenticación'}), 500

    for r in (_ACCESOS_CACHE['rows'] or []):
        r_emp = str(r.get('Empleado', '') or '').strip()
        r_pw  = str(r.get('Contraseña', '') or '').strip()
        r_rol = str(r.get('Rol', '') or 'admin').strip() or 'admin'
        if r_emp == emp and r_pw == pw:
            return jsonify({'ok': True, 'rol': r_rol})
    return jsonify({'ok': False, 'error': 'Credenciales incorrectas'}), 401


# ── PATCH /api/portafolio/proyecto/fase ───────────────────────────────

VALID_FASES = {'Desarrollo', 'Pruebas QA', 'Liberación', 'Estabilización', 'Hecho', 'Cerrado'}

@app.route('/api/portafolio/proyecto/fase', methods=['PATCH'])
def patch_proyecto_fase():
    data  = request.get_json(silent=True) or {}
    folio = str(data.get('folio', '') or '').strip()
    fase  = str(data.get('fase',  '') or '').strip()
    emp   = str(data.get('emp',   '') or '').strip()

    if not folio or fase not in VALID_FASES:
        return jsonify({'error': 'Datos inválidos'}), 400

    wb = get_master_workbook()
    try:
        ws = wb.worksheet('Fases')
    except gspread.exceptions.WorksheetNotFound:
        ws = wb.add_worksheet(title='Fases', rows=200, cols=4)
        ws.append_row(['ID_Proyecto', 'Fase_Actual', 'Actualizado_Em', 'Fecha_Actualizacion'],
                      value_input_option='RAW')

    all_vals = ws.get_all_values()
    headers  = all_vals[0] if all_vals else []
    col_map  = {h: i for i, h in enumerate(headers)}
    ci_folio = col_map.get('ID_Proyecto', 0)
    ci_fase  = col_map.get('Fase_Actual', 1)
    ci_emp   = col_map.get('Actualizado_Em', 2)
    ci_fecha = col_map.get('Fecha_Actualizacion', 3)

    now_str = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')

    for row_i, row in enumerate(all_vals[1:], start=2):
        if len(row) > ci_folio and row[ci_folio].strip() == folio:
            ws.update_acell(col_index_to_a1(ci_fase)  + str(row_i), fase)
            ws.update_acell(col_index_to_a1(ci_emp)   + str(row_i), emp)
            ws.update_acell(col_index_to_a1(ci_fecha) + str(row_i), now_str)
            _PORTAFOLIO_CACHE.clear()  # forzar rebuild para que la fase persista al recargar
            return jsonify({'ok': True, 'updated': True})

    ws.append_row([folio, fase, emp, now_str], value_input_option='USER_ENTERED')
    _PORTAFOLIO_CACHE.clear()  # forzar rebuild para que la fase persista al recargar
    return jsonify({'ok': True, 'created': True})


# ── Main ──────────────────────────────────────────────────────────────

if __name__ == '__main__':
    if not MASTER_SPREADSHEET_ID or not GOOGLE_CREDENTIALS_JSON:
        print('=' * 60)
        print('ATENCION: MASTER_SPREADSHEET_ID o GOOGLE_CREDENTIALS_JSON no configurado.')
        print('Copia .env.example -> .env y completa las variables.')
        print('=' * 60)
    app.run(debug=True, port=5050, host='127.0.0.1')
